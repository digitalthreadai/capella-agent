#!/usr/bin/env python3
"""
requirements_to_excel.py  —  Generic JSON → Excel converter for requirements datasets.

Usage:
    python requirements_to_excel.py <input.json> [output.xlsx]

    If output.xlsx is omitted, the name is derived from the input file.
    Pass '-' as input to read JSON from stdin.

Input format:
    A JSON array of objects.  Every object should have the same keys;
    the first object's keys become column headers in the order they appear.

    Example:
        [
          {
            "Req ID":             "REQ-001",
            "Title":              "Boot time",
            "Statement":          "The system shall boot in under 30 s.",
            "Category":           "Performance",
            "Priority":           "High",
            "Capella Layer":      "System",
            "Traced Element Type":"SystemFunction"
          },
          ...
        ]

Optional conventions (auto-detected, case-insensitive):
    - A column whose name contains "priority" triggers row colour-coding:
          High   → light red    (#FFE0E0)
          Medium → light yellow (#FFF9CC)
          Low    → light green  (#E0F5E0)
    - A column whose name contains "category" triggers row-stripe colouring
      (each distinct value gets a stable pastel colour from a built-in palette).

Output:
    A formatted .xlsx file with:
        - Dark blue header row (white bold text)
        - Auto-sized column widths (capped at 60 chars)
        - Thin borders on all cells
        - Text wrap on columns wider than 40 chars
        - Frozen top row

Dependencies:
    pip install openpyxl
"""

import json
import sys
import os
import argparse
from collections import OrderedDict

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit(
        "ERROR: openpyxl is not installed.\n"
        "Run:  pip install openpyxl\n"
    )

# ── Colour constants ────────────────────────────────────────────────────────

HEADER_BG   = "1F3864"   # dark navy
HEADER_FG   = "FFFFFF"   # white

PRIORITY_COLOURS = {
    "high":   "FFE0E0",   # light red
    "medium": "FFF9CC",   # light yellow
    "low":    "E0F5E0",   # light green
}

# Pastel palette for category row-striping (cycles if > 12 categories)
CATEGORY_PALETTE = [
    "BDD7EE",  # light blue
    "E2EFDA",  # light green
    "FFF2CC",  # light yellow
    "FCE4D6",  # light orange / peach
    "F2CEEF",  # light purple
    "FFCCCC",  # light rose
    "D9D9D9",  # light grey
    "DDEBF7",  # pale sky blue
    "F4CCFF",  # lavender
    "CCF5E0",  # mint
    "FFE4B5",  # moccasin
    "E8D5C4",  # warm sand
]

DEFAULT_BG  = "FFFFFF"   # fallback white

# ── Helpers ─────────────────────────────────────────────────────────────────

def _fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)


def _border() -> Border:
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _find_col(headers: list[str], keyword: str) -> int | None:
    """Return 0-based index of first header containing *keyword* (case-insensitive)."""
    kw = keyword.lower()
    for i, h in enumerate(headers):
        if kw in h.lower():
            return i
    return None


def _cell_str(value) -> str:
    """Coerce any JSON value to a clean string (no trailing .0 on ints)."""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


# ── Main converter ───────────────────────────────────────────────────────────

def convert(records: list[dict], output_path: str) -> None:
    if not records:
        sys.exit("ERROR: Input JSON array is empty.")

    headers = list(records[0].keys())

    # Detect optional feature columns
    pri_col = _find_col(headers, "priority")
    cat_col = _find_col(headers, "category")

    # Assign a stable pastel colour to each category value
    cat_colour_map: dict[str, str] = {}
    if cat_col is not None:
        seen_cats: list[str] = list(
            OrderedDict.fromkeys(
                str(r.get(headers[cat_col], "")).strip()
                for r in records
            )
        )
        for idx, cat in enumerate(seen_cats):
            cat_colour_map[cat] = CATEGORY_PALETTE[idx % len(CATEGORY_PALETTE)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    ws.freeze_panes = "A2"   # freeze header row

    # ── Header row ──
    header_font   = Font(bold=True, color=HEADER_FG, size=11)
    header_fill   = _fill(HEADER_BG)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = _border()

    ws.row_dimensions[1].height = 28

    # Track max content width per column (start with header widths)
    col_widths = [len(h) for h in headers]

    # ── Data rows ──
    for row_idx, record in enumerate(records, start=2):
        # Determine row background
        row_bg = DEFAULT_BG

        # Category stripe takes precedence as base
        if cat_col is not None:
            cat_val = str(record.get(headers[cat_col], "")).strip()
            row_bg = cat_colour_map.get(cat_val, DEFAULT_BG)

        # Priority overlay (overrides category stripe for that row — or blend if you prefer)
        if pri_col is not None:
            pri_val = str(record.get(headers[pri_col], "")).strip().lower()
            if pri_val in PRIORITY_COLOURS:
                row_bg = PRIORITY_COLOURS[pri_val]

        row_fill = _fill(row_bg)

        for col_idx, header in enumerate(headers, start=1):
            raw  = record.get(header, "")
            text = _cell_str(raw)

            needs_wrap = len(text) > 40

            cell = ws.cell(row=row_idx, column=col_idx, value=text)
            cell.fill      = row_fill
            cell.border    = _border()
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=needs_wrap,
                horizontal="left" if needs_wrap else "center",
            )

            # Track width
            col_widths[col_idx - 1] = max(
                col_widths[col_idx - 1],
                min(len(text), 60)     # cap at 60 chars
            )

    # ── Column widths ──
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 3  # padding

    # ── Row heights: auto-scale for wrapped cells ──
    for row_idx in range(2, len(records) + 2):
        # Estimate based on longest cell in the row
        max_lines = 1
        for col_idx, header in enumerate(headers, start=1):
            raw   = _cell_str(records[row_idx - 2].get(header, ""))
            width = col_widths[col_idx - 1]
            lines = max(1, len(raw) // max(width, 1)) + 1 if len(raw) > 40 else 1
            max_lines = max(max_lines, lines)
        ws.row_dimensions[row_idx].height = max(18, max_lines * 14)

    wb.save(output_path)
    print(f"OK: Written {len(records)} requirements -> {output_path}")


# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Convert a JSON requirements array to a formatted Excel file."
    )
    parser.add_argument(
        "input",
        help="Path to JSON file (or '-' to read from stdin)",
    )
    parser.add_argument(
        "output",
        nargs="?",
        help="Output .xlsx path (default: <input_stem>.xlsx)",
    )
    args = parser.parse_args()

    # Read input
    if args.input == "-":
        raw = sys.stdin.read()
    else:
        if not os.path.exists(args.input):
            sys.exit(f"ERROR: Input file not found: {args.input}")
        with open(args.input, encoding="utf-8") as f:
            raw = f.read()

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        sys.exit(f"ERROR: Invalid JSON — {e}")

    if not isinstance(data, list):
        sys.exit("ERROR: JSON must be an array (list) of objects at the top level.")

    # Determine output path
    if args.output:
        out_path = args.output
    elif args.input == "-":
        out_path = "requirements.xlsx"
    else:
        stem = os.path.splitext(args.input)[0]
        out_path = stem + ".xlsx"

    convert(data, out_path)


if __name__ == "__main__":
    main()
